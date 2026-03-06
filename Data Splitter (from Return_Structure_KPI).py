import pandas as pd
import numpy as np
import os
from tkinter import filedialog, messagebox
import tkinter as tk

EXCLUDED_SCHOOLS = {"Health and Safety"}


def select_input_file():
    """
    Opens a file dialog to select the input Excel file.

    Returns:
        str: Path to the selected file, or None if cancelled
    """
    root = tk.Tk()
    root.withdraw()  # Hide the main window

    file_path = filedialog.askopenfilename(
        title="Select the Excel file with school-level KPI data",
        filetypes=[
            ("Excel files", "*.xlsx *.xls"),
            ("All files", "*.*")
        ]
    )

    root.destroy()
    return file_path if file_path else None


def select_output_location():
    """
    Opens a file dialog to select where to save the output file.

    Returns:
        str: Path for the output file, or None if cancelled
    """
    root = tk.Tk()
    root.withdraw()  # Hide the main window

    file_path = filedialog.asksaveasfilename(
        title="Save aggregated data as...",
        defaultextension=".xlsx",
        filetypes=[
            ("Excel files", "*.xlsx"),
            ("All files", "*.*")
        ]
    )

    root.destroy()
    return file_path if file_path else None


def clean_numeric_series(series):
    """
    Clean a pandas series to handle comma-formatted numbers and convert to numeric.
    
    Args:
        series: pandas Series with potentially comma-formatted numbers
        
    Returns:
        pandas Series with cleaned numeric values
    """
    # Convert to string and clean commas, dollar signs, etc.
    cleaned = series.astype(str).str.replace(',', '').str.replace('$', '').str.replace('%', '').str.strip()
    
    # Convert to numeric, coercing errors to NaN
    numeric_series = pd.to_numeric(cleaned, errors='coerce')
    
    # Fill NaN with 0
    return numeric_series.fillna(0)


def parse_reporting_date(value):
    """Parse reporting date values using day-first format."""
    if pd.isna(value) or value == "":
        return pd.NaT
    return pd.to_datetime(value, dayfirst=True, errors='coerce')


def format_reporting_date(value):
    """Format parsed reporting dates consistently for outputs."""
    parsed = parse_reporting_date(value)
    if pd.isna(parsed):
        return ""
    return parsed.strftime("%d/%m/%Y")


def load_kpi_sheet(file_path):
    """Load KPI sheet and apply shared cleaning filters."""
    df = pd.read_excel(file_path, sheet_name='Sheet 1 - Return_Structure_KPI')

    if 'School' in df.columns:
        df = df[~df['School'].astype(str).str.strip().isin(EXCLUDED_SCHOOLS)].copy()

    if 'Date' in df.columns:
        df['Date'] = df['Date'].apply(parse_reporting_date)

    return df


def get_available_reporting_periods(file_path):
    """Return sorted list of reporting periods available in KPI source file."""
    df = load_kpi_sheet(file_path)
    if 'Date' not in df.columns:
        return []
    periods = sorted([d for d in df['Date'].dropna().unique()])
    return periods


def create_tooltips_dataframe():
    """Create Question Tooltips dataframe for dashboard compatibility"""
    tooltip_mapping = {
        "% of Written Arrangements Complete": "Written Arrangements determine how a department is enacting a specific University Policy. Where a policy determines what a department must do, the arrangements indicate how a department achieves this. This value indicates how many arrangements a department has, alongside the percentage of these which are approved by the local H&S Committee and are in-date/reviewed.",
        "% Risk Assessments on Register up-to-date": "This number provides an indication of the percentage of risk assessments that are both on the departmental risk assessment register, AND are reviewed/up-to-date.\n\nPlease review the number of assessments provided alongside the % reviewed, as the number will provide an indication on whether the coverage across the department is high enough.",
        "% of Staff Completed UoN H&S Induction": "This number is calculated by comparing the  number of staff (from the UniCore records) who have completed the mandatory Health and Safety Induction course against the number of staff within the Department.\nNote that training data is provided by the H&S Dept, but Departments were asked to request their staff numbers through a UniCore report as the H&S Dept cannot procure this data.",
        "% of Staff Completed UoN Fire Training": "This number is calculated by comparing the  number of staff (from the UniCore records) who have completed the mandatory Fire Training course against the number of staff within the Department.\nNote that training data is provided by the H&S Dept, but Departments were asked to request their staff numbers through a UniCore report as the H&S Dept cannot procure this data.",
        "% of Fire Drills Carried out": "This is calculated by comparing the number of buildings that were due to have a Fire Drill within the reporting period against the number of drills carried out.\n\nThe number provided alongside the percentage should be roughly half the number of buildings each department is responsible for.",
        "% of PEEPS in Place, Reviewed and Controlled": "This value provides an indication of how many PEEPS have both been identified as required and are implemented. These should have been reviewed in good time and have appropriate controls in place to mitigate the risks related to the evacuation of the PEEPs holders.",
        "% of PEEPS that are tested/drilled": "This value provides an indication of how many of the identified PEEPS have been tested to ensure that, upon execution, the affected person can be successfully evacuated,",
        "% of Assets without active A and B defects": "This return provides an indication of the % of Assets that have been assessed by Allianz as not having a defect that requires action to remediate, where an \"A\" defect corresponds to one that poses a serious hazard, and a \"B\" defect requires action, but not immediately.",
        "% of Assets seen to by Allianz": "This return provides an indication of how many assets the department has on the Allianz database compared to how many have been inspected within the time frame in which they were due an inspection.\n\nNote that this return does not include assets outside of the Allianz database, including items such as gas regulators, autoclaves, etc. A system is not currently available at the University to maintain these records.",
        "% of Incidents + Near Missed Investigated": "This return compares the number of Incidents and Near Misses reported in the period alongside the % that have been investigated by the relevant person in the Dept, whether this be the supervisor of the reporting party, or the Health and Safety Assisstants/Coordinators.",
        "% of Inspections Carried out against Monitoring Schedule": "This return provides an overview of how many Health and Safety Inspections were due to be carried out within the Dept compared to the number actually carried out.\n\nThe return indicates both the number of inspections planned, as well as the number of areas highlighted in the department as requiring inspection. The provided percentage is an indication of how many were successfully carried out against the schedule, but the number schedules should be compared against the number of areas requiring inspection as this will determine whether the number of inspections planned is proportionate against the size of the department",
        "% of Leadership Walkarounds Carried out": "This return provides an overview of the percentage of leadership walkarounds scheduled, vs the number completed.\n\nEach department may expect the number of walkarounds to be roughly 1 per quarter, but this will vary depending on the risk profile of the department.",
        "Percentage Coverage of Risk Assessments": "This return provides an indication of what the H&S Coordinator believes the coverage of Risk Assessments is across each department.\n\nCurrently, there is no University-wide Risk Assessment Register, so in most cases this figure will be an estimate.\n\nRisk Assessment coverage should reflect the range of hazards within a department (with the RMS module list as a useful starting point), and whether the department believes it has an appropriate number of Risk Assessments in place for each hazard type, taking into account the number of research groups and activities in their buildings.",
        "% of Training identified in Matrix that is accessible": "This return provides an indication of how much of the training identified in the Department's Training Matrix — which should capture all training requirements based on the activities carried out by each staff member or student — is currently accessible to those who need it.\n\nTraining accessibility should reflect whether the courses identified in the Matrix are:\n\nAvailable in the required format (e-learning, in-person, or blended),\n\nKept up to date and relevant, and\n\nRealistically accessible to the groups who need them (considering scheduling, prerequisites, and availability of spaces).",
        "% of Staff who are in date with all training requirements": "This return is an indication by the H&S Coordinator of what percentage of Staff they believe have received all training they need in order to be competent to carry out their work. This includes all compulsory training via UniCore, all training via Atlus, as well as any other training highlighted on the training matrix."
    }
    
    # Create dataframe with column names as headers (row 0) and tooltips as row 1
    column_names = list(tooltip_mapping.keys())
    tooltip_texts = list(tooltip_mapping.values())
    
    # Create DataFrame with column names as headers and tooltip text as the data
    # Need 2 rows: row 0 (headers), row 1 (tooltip text)
    df = pd.DataFrame([column_names, tooltip_texts], columns=column_names)
    return df


def find_column_by_variations(df, target_column, variations=None):
    """
    Find a column in the dataframe by checking various name variations.
    
    Args:
        df: pandas DataFrame
        target_column: the exact column name to look for first
        variations: list of alternative column names to try
    
    Returns:
        str: the actual column name found, or None if not found
    """
    # First try the exact match
    if target_column in df.columns:
        return target_column
    
    # If variations provided, try them
    if variations:
        for variation in variations:
            if variation in df.columns:
                return variation
    
    # Try some common automatic variations for Leadership Walkarounds
    if 'Leadership walkarounds' in target_column:
        auto_variations = [
            target_column.replace('walkarounds', 'walkrounds'),
            target_column.replace('No of', 'Number of'),
            target_column.replace('No of', 'No. of'),
            target_column.replace('completed', 'carried out'),
            target_column.replace('completed', 'done'),
            # Try specific variations
            'Number of Leadership walkarounds completed',
            'No of Leadership walkarounds completed',
            'Number of Leadership walkrounds completed',
            'No of Leadership walkrounds completed'
        ]
        
        for variation in auto_variations:
            if variation in df.columns:
                # Don't return percentage columns for numerator fields
                if target_column and ('completed' in target_column or 'carried out' in target_column):
                    if not variation.startswith('%') and '% of' not in variation:
                        return variation
                else:
                    return variation
    
    # Also check for partial matches by looking for columns that contain key terms
    if 'Leadership' in target_column and ('completed' in target_column or 'carried out' in target_column):
        for col in df.columns:
            if ('leadership' in col.lower() and 
                ('completed' in col.lower() or 'carried out' in col.lower()) and
                not col.startswith('%') and '% of' not in col):
                return col
    
    return None


def aggregate_school_to_faculty_data(file_path, reporting_period=None):
    """
    Aggregates school-level data to faculty-level data by correctly calculating percentages
    from summed raw numbers rather than averaging existing percentages.

    Args:
        file_path (str): Path to the Excel file containing school-level KPI data

    Returns:
        pd.DataFrame: Faculty-level aggregated data with recalculated percentages
    """

    # Read and pre-clean data
    df = load_kpi_sheet(file_path)

    if reporting_period is not None and 'Date' in df.columns:
        period = parse_reporting_date(reporting_period)
        if not pd.isna(period):
            df = df[df['Date'] == period].copy()
    
    # Debug: Show available Leadership columns (reduced debug output)
    leadership_cols = [col for col in df.columns if 'leadership' in str(col).lower()]
    if leadership_cols:
        print(f"DEBUG - Available Leadership columns: {leadership_cols}")

    # Define the columns that need aggregation and their corresponding raw number columns
    percentage_calculations = [
        {
            'percentage_col': '% of Written Arrangements Complete',
            'numerator_col': 'Number of Arrangements Completed',
            'denominator_col': 'Number of Arrangements'
        },
        {
            'percentage_col': '% Risk Assessments on Register up-to-date',
            'numerator_col': 'Number of Risk Assessments Updated',
            'denominator_col': 'Number of Risk Assessments on Register'
        },
        {
            'percentage_col': '% of Staff Completed UoN H&S Induction',
            'numerator_col': 'No of Staff Completing H&S Training',
            'denominator_col': 'Number of Staff'
        },
        {
            'percentage_col': '% of Staff Completed UoN Fire Training',
            'numerator_col': 'no of Staff Completing Fire Training',
            'denominator_col': 'Number of Staff'
        },
        {
            'percentage_col': '% of Fire Drills Carried out',
            'numerator_col': 'Number of Fire Drills Carried out',
            'denominator_col': 'Number of Buildings Allocated for Fire Drills to be undertaken'
        },
        {
            'percentage_col': '% of PEEPS in Place, Reviewed and Controlled',
            'numerator_col': 'No of PEEPS in place',
            'denominator_col': 'Number of PEEPS Identified'
        },
        {
            'percentage_col': '% of PEEPS that are tested/drilled',
            'numerator_col': 'No of PEEPS rehearsed',
            'denominator_col': 'Number of PEEPS Identified'
        },
        {
            'percentage_col': '% of Assets without active A and B defects',
            'numerator_col': None,  # Special case: 100 - (A&B defects / Total Assets * 100)
            'denominator_col': None
        },
        {
            'percentage_col': '% of Assets seen to by Allianz',
            'numerator_col': None,  # Special case: 100 - (Assets Overdue / Total Assets * 100)
            'denominator_col': None
        },
        {
            'percentage_col': '% of Incidents + Near Missed Investigated',
            'numerator_col': 'No of Investigations Completed for Incidents Reported in Period',
            'denominator_col': None  # Special case: sum of accidents + near misses
        },
        {
            'percentage_col': '% of Inspections Carried out against Monitoring Schedule',
            'numerator_col': 'Number of Inspections carried out against Monitoring Schedule',
            'denominator_col': 'Number of Inspections on Monitoring Schedule'
        },
        {
            'percentage_col': '% of Leadership Walkarounds Carried out',
            'numerator_col': 'Number of Leadership walkarounds completed',
            'denominator_col': 'Number of Leadership walkarounds on Monitoring Schedule'
        }
    ]

    # Columns that should be averaged (already percentages without clear raw number pairs)
    average_columns = [
        'Percentage Coverage of Risk Assessments',
        '% of Training identified in Matrix that is accessible',
        '% of Staff who are in date with all training requirements'
    ]

    # Columns that should be summed (raw numbers)
    sum_columns = [
        'Number of Arrangements',
        'Number of Arrangements Completed',
        'Number of Risk Assessments on Register',
        'Number of Risk Assessments Updated',
        'Number of Staff',
        'No of Staff Completing H&S Training',
        'no of Staff Completing Fire Training',
        'Number of Buildings Allocated for Fire Drills to be undertaken',
        'Number of Fire Drills Carried out',
        'Number of PEEPS Identified',
        'No of PEEPS in place',
        'No of PEEPS rehearsed',
        'Number of BU Owned Assets',
        'Number of Assets Overdue',
        'Number of Assets seen to by Allianz',
        'No of A & B defects',
        'Total Number of Incidents Still Open',
        'Number of Accidents/Illness',
        'Number of Near Misses',
        'No of Investigations Completed for Incidents Reported in Period',
        'Number of Areas Requiring Inspection',
        'Number of Inspections on Monitoring Schedule',
        'Number of Inspections carried out against Monitoring Schedule',
        'Number of Leadership walkarounds on Monitoring Schedule',
        'Number of Leadership walkarounds completed',
        'Number of Medium/High Actions Due for Completetion',
        'Number of Medium/High Priority Actions Completed'
    ]

    # Clean the data - replace empty strings and None with 0 for numeric columns
    numeric_columns = sum_columns + [calc['numerator_col'] for calc in percentage_calculations] + [
        calc['denominator_col'] for calc in percentage_calculations if calc['denominator_col']]

    for col in numeric_columns:
        if col in df.columns:
            df[col] = clean_numeric_series(df[col])
    
    # Also clean percentage columns that are averaged (may contain inconsistent decimal vs percentage values)
    # First, strip % symbols from all columns that might contain percentages
    percentage_columns = average_columns + [calc['percentage_col'] for calc in percentage_calculations]
    
    for col in df.columns:
        if any(pct_col in str(col) for pct_col in ['%', 'Percentage']):
            # Strip % symbols and convert to numeric using comma-aware cleaning
            if df[col].dtype == 'object':
                df[col] = clean_numeric_series(df[col])
    
    for col in average_columns:
        if col in df.columns:
            # Convert to numeric, handling both decimal (0.5) and percentage (50%) formats
            df[col] = clean_numeric_series(df[col])
            # Convert individual decimal values (between 0 and 1) to percentages
            # This handles mixed formats where some values are decimals and others are percentages
            mask = (df[col] > 0) & (df[col] <= 1)
            df.loc[mask, col] = df.loc[mask, col] * 100

    # Group by Faculty and aggregate
    faculty_data = []

    for faculty in df['Faculty'].unique():
        if pd.isna(faculty):
            continue

        faculty_schools = df[df['Faculty'] == faculty]
        aggregated_row = {'Faculty': faculty}
        if 'Date' in df.columns and reporting_period is not None:
            aggregated_row['Date'] = parse_reporting_date(reporting_period)

        # Sum the raw number columns
        for col in sum_columns:
            actual_col = find_column_by_variations(faculty_schools, col)
            if actual_col:
                aggregated_row[col] = faculty_schools[actual_col].sum()
                if actual_col != col and 'Leadership' in col:
                    print(f"DEBUG - Found Leadership column: '{actual_col}' for '{col}'")
            else:
                aggregated_row[col] = 0

        # Average the percentage columns that don't have clear raw number pairs
        for col in average_columns:
            if col in faculty_schools.columns:
                # Only average non-zero values
                valid_values = faculty_schools[col].dropna()
                valid_values = valid_values[valid_values != 0]
                
                if len(valid_values) > 0:
                    # Handle mixed decimal/percentage formats for all average columns
                    normalized_values = []
                    has_decimals = False
                    has_percentages = False
                    
                    for val in valid_values:
                        if val <= 1:
                            normalized_values.append(val * 100)  # Convert decimal to percentage
                            has_decimals = True
                        else:
                            normalized_values.append(val)  # Already a percentage
                            has_percentages = True
                    
                    aggregated_row[col] = sum(normalized_values) / len(normalized_values)
                    
                    # Debug output for columns that had mixed formats
                    if has_decimals and has_percentages:
                        print(f"DEBUG - {col}: Fixed mixed format, result = {aggregated_row[col]:.1f}%")
                else:
                    aggregated_row[col] = 0

        # Calculate percentages from summed raw numbers
        for calc in percentage_calculations:
            numerator_col = calc['numerator_col']
            denominator_col = calc['denominator_col']
            percentage_col = calc['percentage_col']
            
            # Find actual column names using variations
            actual_numerator_col = find_column_by_variations(df, numerator_col) if numerator_col else None
            actual_denominator_col = find_column_by_variations(df, denominator_col) if denominator_col else None

            if denominator_col is None:
                # Special case for incidents (accidents + near misses)
                if percentage_col == '% of Incidents + Near Missed Investigated':
                    total_incidents = (aggregated_row.get('Number of Accidents/Illness', 0) +
                                       aggregated_row.get('Number of Near Misses', 0))
                    numerator = aggregated_row.get(actual_numerator_col, 0) if actual_numerator_col else 0
                    if total_incidents > 0:
                        aggregated_row[percentage_col] = (numerator / total_incidents) * 100
                    else:
                        aggregated_row[percentage_col] = 0
                # Special case for assets without A&B defects
                elif percentage_col == '% of Assets without active A and B defects':
                    total_assets = aggregated_row.get('Number of BU Owned Assets', 0)
                    defects = aggregated_row.get('No of A & B defects', 0)
                    if total_assets > 0:
                        aggregated_row[percentage_col] = 100 - (defects / total_assets * 100)
                    else:
                        aggregated_row[percentage_col] = 0
                # Special case for assets seen to by Allianz
                elif percentage_col == '% of Assets seen to by Allianz':
                    total_assets = aggregated_row.get('Number of BU Owned Assets', 0)
                    overdue_assets = aggregated_row.get('Number of Assets Overdue', 0)
                    if total_assets > 0:
                        aggregated_row[percentage_col] = 100 - (overdue_assets / total_assets * 100)
                    else:
                        aggregated_row[percentage_col] = 0
            else:
                numerator = aggregated_row.get(actual_numerator_col, 0) if actual_numerator_col else 0
                denominator = aggregated_row.get(actual_denominator_col, 0) if actual_denominator_col else 0

                # Debug output for Leadership Walkarounds
                if 'Leadership' in percentage_col:
                    print(f"DEBUG - {aggregated_row.get('Faculty', 'Unknown')}: Leadership walkarounds {numerator}/{denominator} = {(numerator/denominator*100) if denominator > 0 else 0:.1f}%")

                if denominator > 0:
                    aggregated_row[percentage_col] = (numerator / denominator) * 100
                else:
                    aggregated_row[percentage_col] = 0

        faculty_data.append(aggregated_row)

    # Create the faculty-level DataFrame
    faculty_df = pd.DataFrame(faculty_data)

    # Round percentages to 2 decimal places for better readability
    percentage_cols = [calc['percentage_col'] for calc in percentage_calculations] + average_columns
    for col in percentage_cols:
        if col in faculty_df.columns:
            faculty_df[col] = faculty_df[col].round(2)

    # Sort by date then faculty when available
    sort_columns = [col for col in ['Date', 'Faculty'] if col in faculty_df.columns]
    if sort_columns:
        faculty_df = faculty_df.sort_values(sort_columns).reset_index(drop=True)

    if 'Date' in faculty_df.columns:
        faculty_df['Date'] = faculty_df['Date'].apply(format_reporting_date)

    return faculty_df


def aggregate_school_to_university_data(file_path, reporting_period=None):
    """
    Aggregates school-level data to university-level data by correctly calculating percentages
    from summed raw numbers rather than averaging existing percentages.

    Args:
        file_path (str): Path to the Excel file containing school-level KPI data

    Returns:
        pd.DataFrame: University-level aggregated data with recalculated percentages (single row)
    """

    # Read and pre-clean data
    df = load_kpi_sheet(file_path)

    if reporting_period is not None and 'Date' in df.columns:
        period = parse_reporting_date(reporting_period)
        if not pd.isna(period):
            df = df[df['Date'] == period].copy()

    # Define the columns that need aggregation and their corresponding raw number columns
    percentage_calculations = [
        {
            'percentage_col': '% of Written Arrangements Complete',
            'numerator_col': 'Number of Arrangements Completed',
            'denominator_col': 'Number of Arrangements'
        },
        {
            'percentage_col': '% Risk Assessments on Register up-to-date',
            'numerator_col': 'Number of Risk Assessments Updated',
            'denominator_col': 'Number of Risk Assessments on Register'
        },
        {
            'percentage_col': '% of Staff Completed UoN H&S Induction',
            'numerator_col': 'No of Staff Completing H&S Training',
            'denominator_col': 'Number of Staff'
        },
        {
            'percentage_col': '% of Staff Completed UoN Fire Training',
            'numerator_col': 'no of Staff Completing Fire Training',
            'denominator_col': 'Number of Staff'
        },
        {
            'percentage_col': '% of Fire Drills Carried out',
            'numerator_col': 'Number of Fire Drills Carried out',
            'denominator_col': 'Number of Buildings Allocated for Fire Drills to be undertaken'
        },
        {
            'percentage_col': '% of PEEPS in Place, Reviewed and Controlled',
            'numerator_col': 'No of PEEPS in place',
            'denominator_col': 'Number of PEEPS Identified'
        },
        {
            'percentage_col': '% of PEEPS that are tested/drilled',
            'numerator_col': 'No of PEEPS rehearsed',
            'denominator_col': 'Number of PEEPS Identified'
        },
        {
            'percentage_col': '% of Assets without active A and B defects',
            'numerator_col': None,  # Special case: 100 - (A&B defects / Total Assets * 100)
            'denominator_col': None
        },
        {
            'percentage_col': '% of Assets seen to by Allianz',
            'numerator_col': None,  # Special case: 100 - (Assets Overdue / Total Assets * 100)
            'denominator_col': None
        },
        {
            'percentage_col': '% of Incidents + Near Missed Investigated',
            'numerator_col': 'No of Investigations Completed for Incidents Reported in Period',
            'denominator_col': None  # Special case: sum of accidents + near misses
        },
        {
            'percentage_col': '% of Inspections Carried out against Monitoring Schedule',
            'numerator_col': 'Number of Inspections carried out against Monitoring Schedule',
            'denominator_col': 'Number of Inspections on Monitoring Schedule'
        },
        {
            'percentage_col': '% of Leadership Walkarounds Carried out',
            'numerator_col': 'Number of Leadership walkarounds completed',
            'denominator_col': 'Number of Leadership walkarounds on Monitoring Schedule'
        }
    ]

    # Columns that should be averaged (already percentages without clear raw number pairs)
    average_columns = [
        'Percentage Coverage of Risk Assessments',
        '% of Training identified in Matrix that is accessible',
        '% of Staff who are in date with all training requirements'
    ]

    # Columns that should be summed (raw numbers)
    sum_columns = [
        'Number of Arrangements',
        'Number of Arrangements Completed',
        'Number of Risk Assessments on Register',
        'Number of Risk Assessments Updated',
        'Number of Staff',
        'No of Staff Completing H&S Training',
        'no of Staff Completing Fire Training',
        'Number of Buildings Allocated for Fire Drills to be undertaken',
        'Number of Fire Drills Carried out',
        'Number of PEEPS Identified',
        'No of PEEPS in place',
        'No of PEEPS rehearsed',
        'Number of BU Owned Assets',
        'Number of Assets Overdue',
        'Number of Assets seen to by Allianz',
        'No of A & B defects',
        'Total Number of Incidents Still Open',
        'Number of Accidents/Illness',
        'Number of Near Misses',
        'No of Investigations Completed for Incidents Reported in Period',
        'Number of Areas Requiring Inspection',
        'Number of Inspections on Monitoring Schedule',
        'Number of Inspections carried out against Monitoring Schedule',
        'Number of Leadership walkarounds on Monitoring Schedule',
        'Number of Leadership walkarounds completed',
        'Number of Medium/High Actions Due for Completetion',
        'Number of Medium/High Priority Actions Completed'
    ]

    # Clean the data - replace empty strings and None with 0 for numeric columns
    numeric_columns = sum_columns + [calc['numerator_col'] for calc in percentage_calculations] + [
        calc['denominator_col'] for calc in percentage_calculations if calc['denominator_col']]

    for col in numeric_columns:
        if col in df.columns:
            df[col] = clean_numeric_series(df[col])
    
    # Also clean percentage columns that are averaged (may contain inconsistent decimal vs percentage values)
    # First, strip % symbols from all columns that might contain percentages
    percentage_columns = average_columns + [calc['percentage_col'] for calc in percentage_calculations]
    
    for col in df.columns:
        if any(pct_col in str(col) for pct_col in ['%', 'Percentage']):
            # Strip % symbols and convert to numeric using comma-aware cleaning
            if df[col].dtype == 'object':
                df[col] = clean_numeric_series(df[col])
    
    for col in average_columns:
        if col in df.columns:
            # Convert to numeric, handling both decimal (0.5) and percentage (50%) formats
            df[col] = clean_numeric_series(df[col])
            # Convert individual decimal values (between 0 and 1) to percentages
            # This handles mixed formats where some values are decimals and others are percentages
            mask = (df[col] > 0) & (df[col] <= 1)
            df.loc[mask, col] = df.loc[mask, col] * 100

    # Aggregate across entire university (all schools)
    aggregated_row = {'Faculty': 'University'}
    if 'Date' in df.columns and reporting_period is not None:
        aggregated_row['Date'] = parse_reporting_date(reporting_period)

    # Sum the raw number columns
    for col in sum_columns:
        actual_col = find_column_by_variations(df, col)
        if actual_col:
            aggregated_row[col] = df[actual_col].sum()
        else:
            aggregated_row[col] = 0

    # Average the percentage columns that don't have clear raw number pairs
    for col in average_columns:
        if col in df.columns:
            # Only average non-zero values
            valid_values = df[col].dropna()
            valid_values = valid_values[valid_values != 0]
            if len(valid_values) > 0:
                # Handle mixed decimal/percentage formats for all average columns
                normalized_values = []
                has_decimals = False
                has_percentages = False
                
                for val in valid_values:
                    if val <= 1:
                        normalized_values.append(val * 100)  # Convert decimal to percentage
                        has_decimals = True
                    else:
                        normalized_values.append(val)  # Already a percentage
                        has_percentages = True
                
                aggregated_row[col] = sum(normalized_values) / len(normalized_values)
                
                # Debug output for columns that had mixed formats
                if has_decimals and has_percentages:
                    print(f"DEBUG - {col}: Fixed mixed format, result = {aggregated_row[col]:.1f}%")
            else:
                aggregated_row[col] = 0

    # Calculate percentages from summed raw numbers
    for calc in percentage_calculations:
        numerator_col = calc['numerator_col']
        denominator_col = calc['denominator_col']
        percentage_col = calc['percentage_col']
        
        # Find actual column names using variations
        actual_numerator_col = find_column_by_variations(df, numerator_col) if numerator_col else None
        actual_denominator_col = find_column_by_variations(df, denominator_col) if denominator_col else None

        if denominator_col is None:
            # Special case for incidents (accidents + near misses)
            if percentage_col == '% of Incidents + Near Missed Investigated':
                total_incidents = (aggregated_row.get('Number of Accidents/Illness', 0) +
                                   aggregated_row.get('Number of Near Misses', 0))
                numerator = aggregated_row.get(actual_numerator_col, 0) if actual_numerator_col else 0
                if total_incidents > 0:
                    aggregated_row[percentage_col] = (numerator / total_incidents) * 100
                else:
                    aggregated_row[percentage_col] = 0
            # Special case for assets without A&B defects
            elif percentage_col == '% of Assets without active A and B defects':
                total_assets = aggregated_row.get('Number of BU Owned Assets', 0)
                defects = aggregated_row.get('No of A & B defects', 0)
                if total_assets > 0:
                    aggregated_row[percentage_col] = 100 - (defects / total_assets * 100)
                else:
                    aggregated_row[percentage_col] = 0
            # Special case for assets seen to by Allianz
            elif percentage_col == '% of Assets seen to by Allianz':
                total_assets = aggregated_row.get('Number of BU Owned Assets', 0)
                overdue_assets = aggregated_row.get('Number of Assets Overdue', 0)
                if total_assets > 0:
                    aggregated_row[percentage_col] = 100 - (overdue_assets / total_assets * 100)
                else:
                    aggregated_row[percentage_col] = 0
        else:
            numerator = aggregated_row.get(actual_numerator_col, 0) if actual_numerator_col else 0
            denominator = aggregated_row.get(actual_denominator_col, 0) if actual_denominator_col else 0

            # Debug output for Leadership Walkarounds
            if 'Leadership' in percentage_col:
                print(f"DEBUG - University: Leadership walkarounds {numerator}/{denominator} = {(numerator/denominator*100) if denominator > 0 else 0:.1f}%")

            if denominator > 0:
                aggregated_row[percentage_col] = (numerator / denominator) * 100
            else:
                aggregated_row[percentage_col] = 0

    # Create the university-level DataFrame (single row)
    university_df = pd.DataFrame([aggregated_row])

    # Round percentages to 2 decimal places for better readability
    percentage_cols = [calc['percentage_col'] for calc in percentage_calculations] + average_columns
    for col in percentage_cols:
        if col in university_df.columns:
            university_df[col] = university_df[col].round(2)

    if 'Date' in university_df.columns:
        university_df['Date'] = university_df['Date'].apply(format_reporting_date)

    return university_df


def filter_and_order_columns(df, use_school=False, include_date=False):
    """
    Filters and orders DataFrame columns to match the exact specification.
    
    Args:
        df (pd.DataFrame): DataFrame to filter and reorder
        use_school (bool): If True, use 'School' instead of 'Faculty' as first column
        
    Returns:
        pd.DataFrame: DataFrame with only specified columns in exact order
    """
    # Exact column order as specified
    first_column = 'School' if use_school else 'Faculty'
    required_columns = [first_column]
    if include_date:
        required_columns.append('Date')

    required_columns.extend([
        'Number of Arrangements',
        '% of Written Arrangements Complete',
        'Number of Risk Assessments on Register',
        '% Risk Assessments on Register up-to-date',
        'Number of Staff',
        '% of Staff Completed UoN H&S Induction',
        '% of Staff Completed UoN Fire Training',
        'Number of Buildings Allocated for Fire Drills to be undertaken',
        '% of Fire Drills Carried out',
        'Number of PEEPS Identified',
        '% of PEEPS in Place, Reviewed and Controlled',
        '% of PEEPS that are tested/drilled',
        'Number of BU Owned Assets',
        '% of Assets without active A and B defects',
        '% of Assets seen to by Allianz',
        'Total Incidents (Accidents + Near Misses)',
        '% of Incidents + Near Missed Investigated',
        'Number of Inspections on Monitoring Schedule',
        '% of Inspections Carried out against Monitoring Schedule',
        'Number of Leadership walkarounds on Monitoring Schedule',
        '% of Leadership Walkarounds Carried out',
        'Percentage Coverage of Risk Assessments',
        '% of Training identified in Matrix that is accessible',
        '% of Staff who are in date with all training requirements'
    ])
    
    # Filter to only include columns that exist in the dataframe
    available_columns = [col for col in required_columns if col in df.columns]
    
    # Create filtered dataframe
    filtered_df = df[available_columns].copy()
    
    # For school raw data, use "/" only for complete non-returns, 0 for partial data
    if use_school:
        # Convert numeric columns to object dtype to allow mixed data types
        numeric_kpi_columns = [col for col in filtered_df.columns if col not in ('School', 'Date') and
                              ('Number of' in col or '% of' in col or 'Percentage' in col or 'Total' in col)]
        
        for col in numeric_kpi_columns:
            filtered_df[col] = filtered_df[col].astype('object')
        
        # For each row, check if it's a complete non-return
        for idx, row in filtered_df.iterrows():
            is_complete_non_return = True
            for col in numeric_kpi_columns:
                value = row[col]
                # If any numeric value exists and is not NaN/empty, it's not a complete non-return
                if pd.notna(value) and value != "" and value != 0:
                    is_complete_non_return = False
                    break
            
            if is_complete_non_return:
                # Complete non-return: use "/" for all NaN/empty values
                for col in filtered_df.columns:
                    if pd.isna(filtered_df.at[idx, col]) or filtered_df.at[idx, col] == "":
                        filtered_df.at[idx, col] = "/"
            else:
                # Partial data: use 0 for NaN/empty numeric values, preserve text columns
                for col in filtered_df.columns:
                    if col not in ('School', 'Date'):  # Don't modify school names or date labels
                        if pd.isna(filtered_df.at[idx, col]) or filtered_df.at[idx, col] == "":
                            filtered_df.at[idx, col] = 0
    
    return filtered_df


def create_school_raw_data(file_path, reporting_period=None, include_date=False):
    """
    Extracts raw school-level data and filters to specified columns.
    
    Args:
        file_path (str): Path to the Excel file containing school-level KPI data
        
    Returns:
        pd.DataFrame: School-level data with specified columns only
    """
    # Read and pre-clean data
    df = load_kpi_sheet(file_path)

    if reporting_period is not None and 'Date' in df.columns:
        period = parse_reporting_date(reporting_period)
        if not pd.isna(period):
            df = df[df['Date'] == period].copy()
    
    # Clean percentage data - strip % symbols and normalize formats
    for col in df.columns:
        if any(pct_col in str(col) for pct_col in ['%', 'Percentage']):
            # Strip % symbols and convert to numeric using comma-aware cleaning
            if df[col].dtype == 'object':
                df[col] = clean_numeric_series(df[col])
    
    # Normalize mixed decimal/percentage formats for percentage columns
    percentage_columns = [col for col in df.columns if any(pct_col in str(col) for pct_col in ['%', 'Percentage'])]
    
    for col in percentage_columns:
        if col in df.columns:
            # Convert to numeric and handle mixed formats
            df[col] = clean_numeric_series(df[col])
            
            # For each value, convert decimals to percentages
            mask = (df[col] <= 1) & (df[col] > 0)  # Values between 0 and 1 (exclusive of 0)
            if mask.any():
                print(f"DEBUG - School Raw Data: Normalized {mask.sum()} decimal values in '{col}' column")
            df.loc[mask, col] = df.loc[mask, col] * 100
    
    # Add the calculated Total Incidents column
    accidents_col = df.get('Number of Accidents/Illness', pd.Series(0, index=df.index)).fillna(0)
    near_misses_col = df.get('Number of Near Misses', pd.Series(0, index=df.index)).fillna(0)
    df['Total Incidents (Accidents + Near Misses)'] = accidents_col + near_misses_col
    
    if 'Date' in df.columns:
        df['Date'] = df['Date'].apply(format_reporting_date)

    # Filter and order columns, using School instead of Faculty
    return filter_and_order_columns(df, use_school=True, include_date=include_date)


def create_faculty_summary_table(file_path, output_path=None):
    """
    Creates a summary table with key percentage columns AND their corresponding raw numbers for each faculty.

    Args:
        file_path (str): Path to the Excel file containing school-level KPI data
        output_path (str, optional): Path to save the output Excel file

    Returns:
        pd.DataFrame: Faculty summary table with key percentages and raw numbers
    """

    try:
        # Get the full aggregated data
        full_data = aggregate_school_to_faculty_data(file_path)

        # Define the key columns for the summary with their raw number counterparts
        summary_column_groups = [
            # Faculty name
            ['Faculty'],

            # Written Arrangements
            ['Number of Arrangements', 'Number of Arrangements Completed', '% of Written Arrangements Complete'],

            # Risk Assessments
            ['Number of Risk Assessments on Register', 'Number of Risk Assessments Updated',
             '% Risk Assessments on Register up-to-date'],

            # Staff Training
            ['Number of Staff', 'No of Staff Completing H&S Training', '% of Staff Completed UoN H&S Induction'],

            # Fire Training
            ['Number of Staff', 'no of Staff Completing Fire Training', '% of Staff Completed UoN Fire Training'],

            # Fire Drills
            ['Number of Buildings Allocated for Fire Drills to be undertaken', 'Number of Fire Drills Carried out',
             '% of Fire Drills Carried out'],

            # PEEPs
            ['Number of PEEPS Identified', 'No of PEEPS in place', '% of PEEPS in Place, Reviewed and Controlled'],
            ['Number of PEEPS Identified', 'No of PEEPS rehearsed', '% of PEEPS that are tested/drilled'],

            # Assets
            ['Number of BU Owned Assets', 'Number of Assets Overdue', 'No of A & B defects',
             '% of Assets without active A and B defects', '% of Assets seen to by Allianz'],

            # Incidents (special case - need to calculate total incidents)
            ['Total Incidents (Accidents + Near Misses)',
             'No of Investigations Completed for Incidents Reported in Period',
             '% of Incidents + Near Missed Investigated'],

            # Inspections
            ['Number of Inspections on Monitoring Schedule',
             'Number of Inspections carried out against Monitoring Schedule',
             '% of Inspections Carried out against Monitoring Schedule'],

            # Leadership Walkarounds
            ['Number of Leadership walkarounds on Monitoring Schedule',
             'Number of Leadership walkarounds completed', '% of Leadership Walkarounds Carried out'],

            # Other percentages (averaged, no clear raw numbers)
            ['Percentage Coverage of Risk Assessments'],
            ['% of Training identified in Matrix that is accessible'],
            ['% of Staff who are in date with all training requirements']
        ]

        if 'Date' in full_data.columns:
            summary_column_groups.insert(1, ['Date'])

        # Create a copy of full_data and add the special calculated column
        summary_df = full_data.copy()

        # Add the total incidents column (sum of accidents and near misses)
        accidents_col = summary_df.get('Number of Accidents/Illness', pd.Series(0, index=summary_df.index)).fillna(0)
        near_misses_col = summary_df.get('Number of Near Misses', pd.Series(0, index=summary_df.index)).fillna(0)
        summary_df['Total Incidents (Accidents + Near Misses)'] = accidents_col + near_misses_col

        # Flatten the column groups and filter to only include available columns
        all_summary_columns = []
        for group in summary_column_groups:
            for col in group:
                if col in summary_df.columns and col not in all_summary_columns:
                    all_summary_columns.append(col)

        # Create the final summary with selected columns
        final_summary = summary_df[all_summary_columns].copy()

        # Save to Excel if output path is provided
        if output_path:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Summary sheet with organized columns
                final_summary.to_excel(writer, sheet_name='Faculty_Summary_With_Numbers', index=False)

                # Full data sheet with all columns
                full_data.to_excel(writer, sheet_name='Faculty_Full_Data', index=False)

            print(f"Faculty aggregated data saved to: {output_path}")

        return final_summary

    except Exception as e:
        print(f"Error in create_faculty_summary_table: {e}")
        raise


def create_university_summary_table(file_path, output_path=None):
    """
    Creates a summary table with key percentage columns AND their corresponding raw numbers for university-wide data.

    Args:
        file_path (str): Path to the Excel file containing school-level KPI data
        output_path (str, optional): Path to save the output Excel file

    Returns:
        pd.DataFrame: University summary table with key percentages and raw numbers (single row)
    """

    try:
        # Get the full aggregated data
        full_data = aggregate_school_to_university_data(file_path)

        # Define the key columns for the summary with their raw number counterparts
        summary_column_groups = [
            # Faculty name (will show "University")
            ['Faculty'],

            # Written Arrangements
            ['Number of Arrangements', 'Number of Arrangements Completed', '% of Written Arrangements Complete'],

            # Risk Assessments
            ['Number of Risk Assessments on Register', 'Number of Risk Assessments Updated',
             '% Risk Assessments on Register up-to-date'],

            # Staff Training
            ['Number of Staff', 'No of Staff Completing H&S Training', '% of Staff Completed UoN H&S Induction'],

            # Fire Training
            ['Number of Staff', 'no of Staff Completing Fire Training', '% of Staff Completed UoN Fire Training'],

            # Fire Drills
            ['Number of Buildings Allocated for Fire Drills to be undertaken', 'Number of Fire Drills Carried out',
             '% of Fire Drills Carried out'],

            # PEEPs
            ['Number of PEEPS Identified', 'No of PEEPS in place', '% of PEEPS in Place, Reviewed and Controlled'],
            ['Number of PEEPS Identified', 'No of PEEPS rehearsed', '% of PEEPS that are tested/drilled'],

            # Assets
            ['Number of BU Owned Assets', 'Number of Assets Overdue', 'No of A & B defects',
             '% of Assets without active A and B defects', '% of Assets seen to by Allianz'],

            # Incidents (special case - need to calculate total incidents)
            ['Total Incidents (Accidents + Near Misses)',
             'No of Investigations Completed for Incidents Reported in Period',
             '% of Incidents + Near Missed Investigated'],

            # Inspections
            ['Number of Inspections on Monitoring Schedule',
             'Number of Inspections carried out against Monitoring Schedule',
             '% of Inspections Carried out against Monitoring Schedule'],

            # Leadership Walkarounds
            ['Number of Leadership walkarounds on Monitoring Schedule',
             'Number of Leadership walkarounds completed', '% of Leadership Walkarounds Carried out'],

            # Other percentages (averaged, no clear raw numbers)
            ['Percentage Coverage of Risk Assessments'],
            ['% of Training identified in Matrix that is accessible'],
            ['% of Staff who are in date with all training requirements']
        ]

        if 'Date' in full_data.columns:
            summary_column_groups.insert(1, ['Date'])

        # Create a copy of full_data and add the special calculated column
        summary_df = full_data.copy()

        # Add the total incidents column (sum of accidents and near misses)
        accidents_col = summary_df.get('Number of Accidents/Illness', pd.Series(0, index=summary_df.index)).fillna(0)
        near_misses_col = summary_df.get('Number of Near Misses', pd.Series(0, index=summary_df.index)).fillna(0)
        summary_df['Total Incidents (Accidents + Near Misses)'] = accidents_col + near_misses_col

        # Flatten the column groups and filter to only include available columns
        all_summary_columns = []
        for group in summary_column_groups:
            for col in group:
                if col in summary_df.columns and col not in all_summary_columns:
                    all_summary_columns.append(col)

        # Create the final summary with selected columns
        final_summary = summary_df[all_summary_columns].copy()

        # Save to Excel if output path is provided
        if output_path:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Summary sheet with organized columns
                final_summary.to_excel(writer, sheet_name='University_Summary_With_Numbers', index=False)

                # Full data sheet with all columns
                full_data.to_excel(writer, sheet_name='University_Full_Data', index=False)

            print(f"University aggregated data saved to: {output_path}")

        return final_summary

    except Exception as e:
        print(f"Error in create_university_summary_table: {e}")
        raise


def create_formatted_summary_sheet(writer, data, column_groups):
    """
    Creates a nicely formatted summary sheet with grouped columns.
    """
    try:
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows

        # Create a new worksheet for the formatted summary
        workbook = writer.book
        worksheet = workbook.create_sheet('Faculty_Summary_Formatted')

        # Write headers with grouping
        current_col = 1
        row = 1

        # Write faculty header
        worksheet.cell(row=row, column=current_col, value='Faculty')
        worksheet.cell(row=row, column=current_col).font = Font(bold=True)
        current_col += 1

        # Define colors for different metric groups
        colors = [
            'E8F4FD',  # Light blue
            'FFF2CC',  # Light yellow
            'E1D5E7',  # Light purple
            'D5E8D4',  # Light green
            'FCE5CD',  # Light orange
            'F8CECC',  # Light red
        ]

        color_index = 0

        for group in column_groups[1:]:  # Skip the Faculty column as it's already added
            if len(group) > 1:  # Only for groups with multiple columns
                # Merge cells for group header if needed
                start_col = current_col

                for col_name in group:
                    if col_name in data.columns:
                        worksheet.cell(row=row, column=current_col, value=col_name)
                        cell = worksheet.cell(row=row, column=current_col)
                        cell.font = Font(bold=True)

                        # Apply background color for grouped metrics
                        if len(group) > 1:
                            cell.fill = PatternFill(start_color=colors[color_index % len(colors)],
                                                    end_color=colors[color_index % len(colors)],
                                                    fill_type='solid')

                        current_col += 1

                if len([col for col in group if col in data.columns]) > 1:
                    color_index += 1
            else:
                # Single column group
                col_name = group[0]
                if col_name in data.columns:
                    worksheet.cell(row=row, column=current_col, value=col_name)
                    worksheet.cell(row=row, column=current_col).font = Font(bold=True)
                    current_col += 1

        # Write data rows
        for idx, row_data in data.iterrows():
            row += 1
            col = 1
            for column in data.columns:
                value = row_data[column]
                worksheet.cell(row=row, column=col, value=value)

                # Format percentage columns
                if '%' in str(column) and pd.notna(value):
                    worksheet.cell(row=row, column=col).number_format = '0.00"%"'

                col += 1

        # Adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width

    except ImportError:
        print("Note: openpyxl styling not available. Basic formatting applied.")
    except Exception as e:
        print(f"Warning: Could not apply advanced formatting: {e}")


def main():
    """
    Main function that handles user interaction and file processing.
    """
    print("KPI Aggregation Tool")
    print("=" * 50)
    print("This tool aggregates school-level KPI data to both faculty and university levels")
    print("by correctly calculating percentages from summed raw numbers.\n")

    # Check dependencies first
    try:
        import pandas as pd
        import openpyxl
    except ImportError as e:
        print(f"Missing required dependency: {e}")
        print("Please install missing packages using:")
        print("pip install pandas openpyxl")
        input("\nPress Enter to exit...")
        return

    # Select input file
    print("Step 1: Select the input Excel file...")
    input_file = select_input_file()

    if not input_file:
        print("No input file selected. Exiting.")
        return

    if not os.path.exists(input_file):
        print(f"Error: File '{input_file}' does not exist.")
        return

    print(f"Selected input file: {input_file}")

    # Select output location
    print("\nStep 2: Select where to save the results...")
    output_file = select_output_location()

    if not output_file:
        print("No output location selected. Results will be displayed only.")
        output_file = None
    else:
        print(f"Selected output file: {output_file}")

    print("\nStep 3: Processing data...")

    try:
        periods = get_available_reporting_periods(input_file)
        if not periods:
            raise ValueError("No reporting periods found in source file.")

        latest_period = periods[-1]
        latest_period_label = format_reporting_date(latest_period)
        period_labels = ", ".join(format_reporting_date(p) for p in periods)
        print(f"Detected reporting periods: {period_labels}")
        print(f"Latest reporting period: {latest_period_label}")

        # Create summaries for the latest period (for dashboard compatibility)
        print("Creating Faculty level aggregation...")
        faculty_summary = aggregate_school_to_faculty_data(input_file, reporting_period=latest_period)

        print("Creating University level aggregation...")
        university_summary = aggregate_school_to_university_data(input_file, reporting_period=latest_period)

        # Save latest-period tabs plus multi-period history tabs
        if output_file:
            # Latest-period data for existing dashboards
            faculty_full_data = faculty_summary.copy()
            accidents_col = faculty_full_data.get('Number of Accidents/Illness', pd.Series(0, index=faculty_full_data.index)).fillna(0)
            near_misses_col = faculty_full_data.get('Number of Near Misses', pd.Series(0, index=faculty_full_data.index)).fillna(0)
            faculty_full_data['Total Incidents (Accidents + Near Misses)'] = accidents_col + near_misses_col
            
            university_full_data = university_summary.copy()
            accidents_col_uni = university_full_data.get('Number of Accidents/Illness', pd.Series(0, index=university_full_data.index)).fillna(0)
            near_misses_col_uni = university_full_data.get('Number of Near Misses', pd.Series(0, index=university_full_data.index)).fillna(0)
            university_full_data['Total Incidents (Accidents + Near Misses)'] = accidents_col_uni + near_misses_col_uni
            
            # Get school raw data for latest period
            school_raw_data = create_school_raw_data(input_file, reporting_period=latest_period, include_date=False)

            # Build history across all periods for trend-ready dashboards
            faculty_history_parts = []
            university_history_parts = []
            school_history_parts = []

            for period in periods:
                faculty_period = aggregate_school_to_faculty_data(input_file, reporting_period=period)
                university_period = aggregate_school_to_university_data(input_file, reporting_period=period)
                school_period = create_school_raw_data(input_file, reporting_period=period, include_date=True)

                faculty_history_parts.append(faculty_period)
                university_history_parts.append(university_period)
                school_history_parts.append(school_period)

            faculty_history = pd.concat(faculty_history_parts, ignore_index=True) if faculty_history_parts else pd.DataFrame()
            university_history = pd.concat(university_history_parts, ignore_index=True) if university_history_parts else pd.DataFrame()
            school_history = pd.concat(school_history_parts, ignore_index=True) if school_history_parts else pd.DataFrame()
            
            # Apply column filtering to all datasets
            university_filtered = filter_and_order_columns(university_full_data)
            faculty_filtered = filter_and_order_columns(faculty_full_data)
            university_history_filtered = filter_and_order_columns(university_history, include_date=True)
            faculty_history_filtered = filter_and_order_columns(faculty_history, include_date=True)
            
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # Core tabs (latest period, compatible with existing dashboards)
                university_filtered.to_excel(writer, sheet_name='University_Summary', index=False)
                faculty_filtered.to_excel(writer, sheet_name='Faculty_Summary', index=False)
                school_raw_data.to_excel(writer, sheet_name='School_Raw_Data', index=False)

                # Multi-period history tabs (for progress/trend analytics)
                university_history_filtered.to_excel(writer, sheet_name='University_Summary_History', index=False)
                faculty_history_filtered.to_excel(writer, sheet_name='Faculty_Summary_History', index=False)
                school_history.to_excel(writer, sheet_name='School_Raw_Data_History', index=False)
                
                # Add Question Tooltips sheet
                tooltips_df = create_tooltips_dataframe()
                tooltips_df.to_excel(writer, sheet_name='Question Tooltips', index=False)
            
            print(f"\n✓ Results saved to: {output_file}")
            print("  7 Tabs Created:")
            print(f"  - University_Summary (latest period: {latest_period_label})")
            print(f"  - Faculty_Summary (latest period: {latest_period_label})")
            print(f"  - School_Raw_Data (latest period: {latest_period_label})")
            print("  - University_Summary_History (all periods)")
            print("  - Faculty_Summary_History (all periods)")
            print("  - School_Raw_Data_History (all periods)")
            print("  - Question Tooltips (KPI metric descriptions for dashboards)")

        print("\nStep 4: Results")
        print("=" * 80)
        print("Faculty Summary Table (with Specified Columns):")

        # Add Total Incidents column for display
        accidents_col = faculty_summary.get('Number of Accidents/Illness', pd.Series(0, index=faculty_summary.index)).fillna(0)
        near_misses_col = faculty_summary.get('Number of Near Misses', pd.Series(0, index=faculty_summary.index)).fillna(0)
        faculty_summary['Total Incidents (Accidents + Near Misses)'] = accidents_col + near_misses_col
        
        accidents_col_uni = university_summary.get('Number of Accidents/Illness', pd.Series(0, index=university_summary.index)).fillna(0)
        near_misses_col_uni = university_summary.get('Number of Near Misses', pd.Series(0, index=university_summary.index)).fillna(0)
        university_summary['Total Incidents (Accidents + Near Misses)'] = accidents_col_uni + near_misses_col_uni
        
        # Display faculty results using filtered data
        faculty_display = filter_and_order_columns(faculty_summary)
        for idx, row in faculty_display.iterrows():
            faculty_name = row['Faculty']
            print(f"\n{faculty_name}:")
            print("-" * (len(faculty_name) + 1))

            # Written Arrangements
            if all(col in row for col in ['Number of Arrangements', 'Number of Arrangements Completed',
                                          '% of Written Arrangements Complete']):
                print(
                    f"  Written Arrangements: {row['Number of Arrangements Completed']:.0f}/{row['Number of Arrangements']:.0f} = {row['% of Written Arrangements Complete']:.2f}%")

            # Risk Assessments
            if all(col in row for col in
                   ['Number of Risk Assessments on Register', 'Number of Risk Assessments Updated',
                    '% Risk Assessments on Register up-to-date']):
                print(
                    f"  Risk Assessments: {row['Number of Risk Assessments Updated']:.0f}/{row['Number of Risk Assessments on Register']:.0f} = {row['% Risk Assessments on Register up-to-date']:.2f}%")

            # H&S Training
            if all(col in row for col in ['Number of Staff', 'No of Staff Completing H&S Training',
                                          '% of Staff Completed UoN H&S Induction']):
                print(
                    f"  H&S Training: {row['No of Staff Completing H&S Training']:.0f}/{row['Number of Staff']:.0f} = {row['% of Staff Completed UoN H&S Induction']:.2f}%")

            # Fire Training
            if all(col in row for col in ['Number of Staff', 'no of Staff Completing Fire Training',
                                          '% of Staff Completed UoN Fire Training']):
                print(
                    f"  Fire Training: {row['no of Staff Completing Fire Training']:.0f}/{row['Number of Staff']:.0f} = {row['% of Staff Completed UoN Fire Training']:.2f}%")

        # Display faculty statistics
        print(f"\n" + "=" * 80)
        print(f"Faculty Summary Statistics:")
        print(f"- Total Faculties Processed: {len(faculty_summary)}")
        print(f"- Faculties: {', '.join(faculty_summary['Faculty'].tolist())}")

        # Display university results
        print(f"\n" + "=" * 80)
        print("University Summary Table (with Specified Columns):")

        # Use filtered data for display  
        university_display = filter_and_order_columns(university_summary)
        row = university_display.iloc[0]
        print(f"\nUniversity Totals:")
        print("-" * 17)

        # Written Arrangements
        if all(col in row for col in ['Number of Arrangements', 'Number of Arrangements Completed',
                                      '% of Written Arrangements Complete']):
            print(
                f"  Written Arrangements: {row['Number of Arrangements Completed']:.0f}/{row['Number of Arrangements']:.0f} = {row['% of Written Arrangements Complete']:.2f}%")

        # Risk Assessments
        if all(col in row for col in
               ['Number of Risk Assessments on Register', 'Number of Risk Assessments Updated',
                '% Risk Assessments on Register up-to-date']):
            print(
                f"  Risk Assessments: {row['Number of Risk Assessments Updated']:.0f}/{row['Number of Risk Assessments on Register']:.0f} = {row['% Risk Assessments on Register up-to-date']:.2f}%")

        # H&S Training
        if all(col in row for col in ['Number of Staff', 'No of Staff Completing H&S Training',
                                      '% of Staff Completed UoN H&S Induction']):
            print(
                f"  H&S Training: {row['No of Staff Completing H&S Training']:.0f}/{row['Number of Staff']:.0f} = {row['% of Staff Completed UoN H&S Induction']:.2f}%")

        # Fire Training
        if all(col in row for col in ['Number of Staff', 'no of Staff Completing Fire Training',
                                      '% of Staff Completed UoN Fire Training']):
            print(
                f"  Fire Training: {row['no of Staff Completing Fire Training']:.0f}/{row['Number of Staff']:.0f} = {row['% of Staff Completed UoN Fire Training']:.2f}%")

        print(f"\n" + "=" * 80)
        print("Both Faculty and University aggregations complete!")

        input("\nPress Enter to exit...")

    except FileNotFoundError as e:
        print(f"Error: Could not find the specified file: {e}")
        print("Please ensure the file path is correct and the file exists.")
    except Exception as e:
        print(f"Error processing file: {e}")
        print("Please ensure the file structure matches the expected format.")
        print("\nExpected format:")
        print("- Excel file with a sheet containing school-level KPI data")
        print("- Columns: 'School', 'Faculty', and various KPI columns")
        input("\nPress Enter to exit...")


# Alternative function for direct usage (when file paths are known)
def process_known_files(input_file_path, output_file_path):
    """
    Process files when paths are already known (for scripting purposes).

    Args:
        input_file_path (str): Path to input Excel file
        output_file_path (str): Path for output Excel file

    Returns:
        pd.DataFrame: Faculty summary table
    """
    if not os.path.exists(input_file_path):
        raise FileNotFoundError(f"Input file not found: {input_file_path}")

    return create_faculty_summary_table(input_file_path, output_file_path)


# Example usage
if __name__ == "__main__":
    # Option 1: Interactive mode with file dialogs (recommended for most users)
    main()

    # Option 2: Direct file paths (uncomment and modify if you know the exact paths)
    # try:
    #     input_file = "path/to/your/Return_Structure_KPI.xlsx"
    #     output_file = "path/to/your/Faculty_Aggregated_KPIs.xlsx"
    #     
    #     faculty_summary = process_known_files(input_file, output_file)
    #     print("Faculty Summary Table:")
    #     print("=" * 80)
    #     print(faculty_summary.to_string(index=False))
    #     
    # except Exception as e:
    #     print(f"Error: {e}")
