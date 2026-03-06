import pandas as pd
import numpy as np
import os
from pathlib import Path


class KPIDataTransformer:
    def __init__(self):
        self.training_percentage_cutover = pd.Timestamp("2026-02-01")
        self.excluded_schools = {"Health and Safety"}
        self.faculty_school_mapping = self._define_faculty_school_mapping()
        self.tooltip_mapping = self._define_tooltip_mapping()
        
    def _define_faculty_school_mapping(self):
        """Define the exact mapping between schools and faculties"""
        return {
            # Arts Faculty
            "CLAS": "Arts",
            "English": "Arts", 
            "Humanities": "Arts",

            # Engineering Faculty
            "Engineering": "Engineering",

            # Estates Faculty
            "Estates": "Estates",

            # Finance and Infrastructure Faculty
            "Finance and Infrastructure": "Finance and Infrastructure",

            # HR Faculty
            "HR": "HR",

            # Medicine & Health Sciences Faculty
            "BDI": "Medicine & Health Sciences",
            "Life Sciences": "Medicine & Health Sciences",
            "Medicine": "Medicine & Health Sciences",
            "Veterinary Medicine and Science": "Medicine & Health Sciences",
            "Clinical Skills": "Medicine & Health Sciences",
            "Health Sciences": "Medicine & Health Sciences",

            # Registrars Faculty
            "Libraries": "Registrars",
            "Sport": "Registrars", 
            "BSU": "Registrars",
            "Student & Public Facing Services": "Registrars",

            # Science Faculty
            "Computer Sciences": "Science",
            "Biosciences": "Science",
            "Chemistry": "Science",
            "Mathematical Sciences": "Science", 
            "Pharmacy": "Science",
            "Physics and Astronomy": "Science",
            "Psychology": "Science",

            # Social Sciences Faculty
            "Economics": "Social Sciences",
            "Business School": "Social Sciences",
            "Education": "Social Sciences",
            "Law": "Social Sciences",
            "Politics and IR": "Social Sciences", 
            "Rights Lab": "Social Sciences",
            "Sociology": "Social Sciences",
            "Geography": "Social Sciences",
            "Faculty Office": "Social Sciences"
        }
        
    def _define_tooltip_mapping(self):
        """Define tooltip mappings for KPI metrics"""
        return {
            "% of Written Arrangements Complete": "Written Arrangements determine how a department is enacting a specific University Policy. A Good/Green Arrangement will describe; the specific University Policy the BU is aiming to comply with, how it is complying (procedures/processes, etc.), what happens when it goes wrong, and how the BU will monitor/validate its approach. They should be regularly reviewed.",
            "% Risk Assessments on Register up-to-date": "This number provides an indication of the percentage of risk assessments that are both on the departmental risk assessment register, AND are reviewed/up-to-date. Risk assessments should be reviewed when something changes, when a recommendation/action was identified, or at least annually.",
            "Percentage Coverage of Risk Assessments": "This is the percentage of work areas that have had a recent (within 18 months) risk assessment against the total number of work areas the BU has identified as requiring a risk assessment.",
            "% of Staff Completed UoN H&S Induction": "This is the percentage of current staff that have completed an UoN specific Health and Safety Induction - this mandatory training introduces staff to UoN specific H&S information.",
            "% of Staff Completed UoN Fire Training": "This is the percentage of current staff that have completed an UoN specific Fire Safety training module.",
            "% of Training identified in Matrix that is accessible": "When a BU identifies staff training requirements (because of specific hazards, legal requirement, etc.), this is the percentage of that training that is accessible to staff either via Learn UoN or via a specific training provider.",
            "% of Staff who are in date with all training requirements": "This is the percentage of current staff who are recorded as in-date with all their identified training requirements.",
            "% of Fire Drills Carried out": "UoN requires that fire drills are carried out in all buildings on an annual basis - this is the percentage that have been carried out as expected.",
            "% of PEEPS in Place, Reviewed and Controlled": "PEEPS (Personal Emergency Evacuation Plans) should be in place for staff/students/visitors who have particular support needs for an evacuation. This is the percentage of identified people who have a PEEP in place which is both up-to-date and effective.",
            "% of PEEPS that are tested/drilled": "PEEPS should be drilled and tested at least annually - this is the percentage that have been.",
            "% of Assets without active A and B defects": "A defects require immediate/urgent attention, B defects should be resolved within 5 years. This is the percentage of BU-owned structures that do not have any A or B defects.",
            "% of Assets seen to by Allianz": "This is the percentage of BU-owned structures that have been seen to/visited by Allianz - this should be 100% as Allianz is contracted to visit all our building stock on a 5-year cycle.",
            "% of Incidents + Near Missed Investigated": "When incidents (accidents/illness) and near misses occur at the university, they should be investigated and the investigation completed in a timely manner. This is the percentage of incidents that have been appropriately investigated.",
            "% of Inspections Carried out against Monitoring Schedule": "BUs should have a monitoring schedule that identifies when/how often specific areas are to be inspected (eg. laboratories inspected monthly). This is the percentage of scheduled inspections that have taken place as expected.",
            "% of Leadership Walkarounds Carried out": "BUs should arrange for their senior staff to undertake Health & Safety Leadership walkarounds - these are planned, meaningful engagement exercises. This is the percentage that have taken place as expected against a monitoring schedule."
        }

    def clean_numeric_value(self, value):
        """Clean and convert a value to float, handling commas and various formats"""
        if pd.isna(value) or value == '' or value is None:
            return np.nan
        
        # Convert to string and clean
        str_value = str(value).strip()
        
        # Remove commas and other formatting
        cleaned_value = str_value.replace(',', '').replace('$', '').replace('%', '')
        
        try:
            return float(cleaned_value)
        except (ValueError, TypeError):
            return np.nan

    def safe_divide(self, numerator, denominator):
        """Safely divide two numbers, handling zero division, NaN values, and comma formatting"""
        # Clean both values
        clean_numerator = self.clean_numeric_value(numerator)
        clean_denominator = self.clean_numeric_value(denominator)
        
        if pd.isna(clean_numerator) or pd.isna(clean_denominator) or clean_denominator == 0:
            return np.nan
        try:
            result = clean_numerator / clean_denominator * 100
            return round(result, 6)  # Round to 6 decimal places to match Excel
        except (ValueError, TypeError, ZeroDivisionError):
            return np.nan

    def safe_subtract_percentage(self, total, defects):
        """Calculate 100 - (defects/total * 100) safely"""
        clean_total = self.clean_numeric_value(total)
        clean_defects = self.clean_numeric_value(defects)
        
        if pd.isna(clean_total) or pd.isna(clean_defects) or clean_total == 0:
            return np.nan
        try:
            percentage = clean_defects / clean_total * 100
            result = 100 - percentage
            return round(result, 6)
        except (ValueError, TypeError, ZeroDivisionError):
            return np.nan

    def safe_incident_calculation(self, investigations, accidents, near_misses):
        """Calculate incidents investigation percentage: investigations / (accidents + near_misses) * 100"""
        clean_investigations = self.clean_numeric_value(investigations)
        clean_accidents = self.clean_numeric_value(accidents)
        clean_near_misses = self.clean_numeric_value(near_misses)
        
        if pd.isna(clean_investigations) or pd.isna(clean_accidents) or pd.isna(clean_near_misses):
            return np.nan
        
        total_incidents = clean_accidents + clean_near_misses
        if total_incidents == 0:
            return np.nan
            
        try:
            result = clean_investigations / total_incidents * 100
            return round(result, 6)
        except (ValueError, TypeError, ZeroDivisionError):
            return np.nan

    def parse_period_date(self, value):
        """Parse report period dates that arrive in UK format."""
        if pd.isna(value) or value == "":
            return None
        parsed = pd.to_datetime(value, dayfirst=True, errors='coerce')
        if pd.isna(parsed):
            return None
        return parsed

    def normalize_percentage_value(self, value):
        """Normalize mixed percentage formats (e.g. 0.75, 75, 75%)."""
        numeric = self.clean_numeric_value(value)
        if pd.isna(numeric):
            return np.nan
        if 0 < numeric <= 1:
            return numeric * 100
        return numeric

    def _create_tooltips_sheet(self, workbook):
        """Create Question Tooltips sheet in the workbook"""
        # Create new worksheet for tooltips
        tooltip_ws = workbook.create_sheet(title="Question Tooltips")
        
        # Row 1: Column names as headers
        for idx, column_name in enumerate(self.tooltip_mapping.keys(), 1):
            tooltip_ws.cell(row=1, column=idx, value=column_name)
        
        # Row 2: Tooltip text
        for idx, tooltip_text in enumerate(self.tooltip_mapping.values(), 1):
            tooltip_ws.cell(row=2, column=idx, value=tooltip_text)

    def transform_data(self, input_csv_path, output_excel_path):
        """Transform Raw Data.csv into Tidyed Data.xlsx format"""
        
        # Read the raw CSV data
        print("Loading Raw Data.csv...")
        raw_df = pd.read_csv(input_csv_path)
        
        # Create the transformed dataframe with headers first
        headers = [
            "School", "Faculty", "Date",
            "Number of Arrangements", "Number of Arrangements Completed", "% of Written Arrangements Complete",
            "Number of Risk Assessments on Register", "Number of Risk Assessments Updated", "% Risk Assessments on Register up-to-date", "Percentage Coverage of Risk Assessments",
            "Number of Staff", "No of Staff Completing H&S Training", "% of Staff Completed UoN H&S Induction", 
            "no of Staff Completing Fire Training", "% of Staff Completed UoN Fire Training",
            "% of Training identified in Matrix that is accessible", "% of Staff who are in date with all training requirements",
            "Number of Buildings Allocated for Fire Drills to be undertaken", "Number of Fire Drills Carried out", "% of Fire Drills Carried out",
            "Number of PEEPS Identified", "No of PEEPS in place", "% of PEEPS in Place, Reviewed and Controlled",
            "No of PEEPS rehearsed", "% of PEEPS that are tested/drilled", 
            "Number of BU Owned Assets", "Number of Assets Overdue", "% of Assets without active A and B defects",
            "No of A & B defects", "% of Assets seen to by Allianz",
            "Total Number of Incidents Still Open", "Number of Accidents/Illness", "Number of Near Misses", 
            "No of Investigations Completed for Incidents Reported in Period", "% of Incidents + Near Missed Investigated",
            "Number of Areas Requiring Inspection", "Number of Inspections on Monitoring Schedule", 
            "Number of Inspections carried out against Monitoring Schedule", "% of Inspections Carried out against Monitoring Schedule",
            "Number of Leadership walkarounds on Monitoring Schedule", "Number of Leadership walkarounds completed", "% of Leadership Walkarounds Carried out",
            "Sources of Actions", "Number of Medium/High Actions Due for Completetion", "Number of Medium/High Priority Actions Completed",
            "Action Log4", "Return_completed?", "BU_arrangements_comments", "RA_comment", "H&S Training_5_confidence",
            "Risk Assessment3_confidence", "H&S_training_comments", "fire_drills_comments", "PEEPS_comments", 
            "assets_comments", "incidents_comments", "monitoring_comments", "actions_comments"
        ]
        
        # Initialize transformed dataframe - add headers as first row
        transformed_data = []
        
        # Add headers as first row of data
        transformed_data.append(headers)
        
        # Exclude entities that should not be reported in dashboards/aggregations.
        if 'School' in raw_df.columns:
            raw_df = raw_df[~raw_df['School'].astype(str).str.strip().isin(self.excluded_schools)].copy()

        print("Processing school data...")
        
        # Process each school
        for idx, row in raw_df.iterrows():
            school_name = row['School']
            faculty_name = self.faculty_school_mapping.get(school_name, "Unknown")
            period_date = self.parse_period_date(row.get('Date', ''))
            training_is_percentage_mode = (
                period_date is not None and period_date >= self.training_percentage_cutover
            )

            staff_count = self.clean_numeric_value(row.get('H&S Training1'))

            if training_is_percentage_mode:
                hs_induction_pct = self.normalize_percentage_value(row.get('H&S Training2'))
                fire_training_pct = self.normalize_percentage_value(row.get('H&S Training3'))

                hs_induction_count = np.nan
                fire_training_count = np.nan
                if pd.notna(staff_count) and staff_count > 0:
                    if pd.notna(hs_induction_pct):
                        hs_induction_count = staff_count * hs_induction_pct / 100.0
                    if pd.notna(fire_training_pct):
                        fire_training_count = staff_count * fire_training_pct / 100.0
            else:
                hs_induction_count = self.clean_numeric_value(row.get('H&S Training2'))
                fire_training_count = self.clean_numeric_value(row.get('H&S Training3'))
                hs_induction_pct = self.safe_divide(hs_induction_count, staff_count)
                fire_training_pct = self.safe_divide(fire_training_count, staff_count)
            
            
            # Build the transformed row
            transformed_row = [
                school_name,  # School
                faculty_name,  # Faculty
                row.get('Date', ''),  # Date
                
                # Arrangements
                row.get('Arrangements1', ''),  # Number of Arrangements
                row.get('Arrangements2', ''),  # Number of Arrangements Completed
                self.safe_divide(row.get('Arrangements2'), row.get('Arrangements1')),  # % Complete
                
                # Risk Assessments
                row.get('Risk Assessment1', ''),  # Number of Risk Assessments on Register
                row.get('Risk Assessment2', ''),  # Number of Risk Assessments Updated
                self.safe_divide(row.get('Risk Assessment2'), row.get('Risk Assessment1')),  # % up-to-date
                row.get('Risk Assessment3', ''),  # Percentage Coverage of Risk Assessments (direct from raw)
                
                # H&S Training
                staff_count,  # Number of Staff
                hs_induction_count,  # No of Staff Completing H&S Training
                hs_induction_pct,  # % H&S Induction
                fire_training_count,  # no of Staff Completing Fire Training
                fire_training_pct,  # % Fire Training
                row.get('H&S Training4', ''),  # % of Training identified in Matrix that is accessible
                row.get('H&S Training5', ''),  # % of Staff who are in date with all training requirements
                
                # Fire Drills
                row.get('Fire Drills1', ''),  # Number of Buildings Allocated for Fire Drills
                row.get('Fire Drills2', ''),  # Number of Fire Drills Carried out
                self.safe_divide(row.get('Fire Drills2'), row.get('Fire Drills1')),  # % Fire Drills Carried out
                
                # PEEPS
                row.get('PEEPS1', ''),  # Number of PEEPS Identified
                row.get('PEEPS2', ''),  # No of PEEPS in place
                self.safe_divide(row.get('PEEPS2'), row.get('PEEPS1')),  # % PEEPS in Place
                row.get('PEEPS3', ''),  # No of PEEPS rehearsed
                self.safe_divide(row.get('PEEPS3'), row.get('PEEPS1')),  # % PEEPS tested/drilled
                
                # BU Owned Assets
                row.get('BU Owned Assets1', ''),  # Number of BU Owned Assets
                row.get('BU Owned Assets2', ''),  # Number of Assets Overdue (A&B defects)
                self.safe_subtract_percentage(row.get('BU Owned Assets1'), row.get('BU Owned Assets2')),  # % without A&B defects
                row.get('BU Owned Assets2', ''),  # No of A & B defects (duplicate for structure)
                self.safe_subtract_percentage(row.get('BU Owned Assets1'), row.get('BU Owned Assets3')),  # % seen by Allianz
                
                # Incidents and Investigations
                row.get('Incidents and Investigations1', ''),  # Total Number of Incidents Still Open
                row.get('Incidents and Investigations2', ''),  # Number of Accidents/Illness
                row.get('Incidents and Investigations3', ''),  # Number of Near Misses
                row.get('Incidents and Investigations4', ''),  # No of Investigations Completed
                self.safe_incident_calculation(
                    row.get('Incidents and Investigations4'), 
                    row.get('Incidents and Investigations2'), 
                    row.get('Incidents and Investigations3')
                ),  # % of Incidents + Near Missed Investigated
                
                # Monitoring Schedule
                row.get('Monitoring Schedule1', ''),  # Number of Areas Requiring Inspection
                row.get('Monitoring Schedule2', ''),  # Number of Inspections on Monitoring Schedule
                row.get('Monitoring Schedule3', ''),  # Number of Inspections carried out
                self.safe_divide(row.get('Monitoring Schedule3'), row.get('Monitoring Schedule2')),  # % Inspections Carried out
                
                # Leadership Walkarounds
                row.get('Leadership walkarounds1', ''),  # Number of Leadership walkarounds on Monitoring Schedule
                row.get('Leadership walkarounds2', ''),  # Number of Leadership walkarounds completed
                self.safe_divide(row.get('Leadership walkarounds2'), row.get('Leadership walkarounds1')),  # % Leadership Walkarounds
                
                # Action Log
                row.get('Action Log1', ''),  # Sources of Actions (mapped to Action Log1 field)
                row.get('Action Log2', ''),  # Number of Medium/High Actions Due for Completion
                row.get('Action Log3', ''),  # Number of Medium/High Priority Actions Completed
                '',  # Action Log4 (empty placeholder to match structure)
                
                # Additional fields from raw data
                row.get('Return_completed?', ''),
                row.get('BU_arrangements_comments', ''),
                row.get('RA_comment', ''),
                row.get('H&S Training_5_confidence', ''),
                row.get('Risk Assessment3_confidence', ''),
                row.get('H&S_training_comments', ''),
                row.get('fire_drills_comments', ''),
                row.get('PEEPS_comments', ''),
                row.get('assets_comments', ''),
                row.get('incidents_comments', ''),
                row.get('monitoring_comments', ''),
                row.get('actions_comments', '')
            ]
            
            transformed_data.append(transformed_row)
        
        print("Creating Excel output...")
        
        # Write to Excel file using openpyxl with manual approach to preserve all rows
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sheet 1 - Return_Structure_KPI'
        
        # Write data row by row to ensure nothing gets dropped
        for row_idx, data_row in enumerate(transformed_data, 1):
            for col_idx, value in enumerate(data_row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Create Question Tooltips sheet
        self._create_tooltips_sheet(wb)
        
        wb.save(output_excel_path)
        
        print(f"✅ Transformation completed successfully!")
        print(f"📊 Input: {len(raw_df)} schools processed")
        print(f"💾 Output: {output_excel_path}")
        print(f"📋 Structure: 58 columns with calculated KPI percentages")
        print(f"💬 Tooltips: Question Tooltips sheet added")

    def run(self, input_csv=None, output_excel=None):
        """Main execution function"""
        
        # Use provided paths or default ones
        if input_csv is None:
            input_csv = "Raw Data.csv"
        if output_excel is None:
            output_excel = "Tidyed Data.xlsx"
            
        # Check if input file exists
        if not os.path.exists(input_csv):
            print(f"❌ Error: Input file '{input_csv}' not found.")
            return False
            
        try:
            self.transform_data(input_csv, output_excel)
            return True
        except Exception as e:
            print(f"❌ Error during transformation: {str(e)}")
            return False


if __name__ == "__main__":
    # Initialize transformer
    transformer = KPIDataTransformer()
    
    # Run transformation 
    print("KPI Data Transformer")
    print("=" * 40)
    print("Converting Raw Data.csv → Tidyed Data.xlsx")
    print()
    
    success = transformer.run()
    
    if success:
        print("\n🎉 Data transformation completed successfully!")
        print("You can now use the Tidyed Data.xlsx with existing dashboard generators.")
    else:
        print("\n💥 Data transformation failed. Please check the error messages above.")
